import zipfile
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os
import logging

logger = logging.getLogger(__name__)


def hex_to_rvb(hex_color: str) -> tuple:
    if hex_color is None:
        return None
    if hex_color.startswith("FF"):
        hex_color = hex_color[2:]
    try:
        r = int(hex_color[0:2], 16)
        v = int(hex_color[2:4], 16)
        b = int(hex_color[4:6], 16)
        return (r, v, b)
    except ValueError:
        logger.warning("Code couleur hex invalide : %s", hex_color, exc_info=True)
        return None


def extract_theme_colors(file_path: str) -> dict:
    theme_colors = {}
    if not os.path.exists(file_path):
        logger.error("Fichier thème introuvable : %s", file_path)
        return theme_colors
    try:
        with zipfile.ZipFile(file_path, 'r') as zip_ref:
            if 'xl/theme/theme1.xml' not in zip_ref.namelist():
                logger.warning("Pas de thème dans le fichier : %s", file_path)
                return theme_colors
            with zip_ref.open('xl/theme/theme1.xml') as theme_file:
                tree = ET.parse(theme_file)
                root = tree.getroot()
                ns = {'a': 'http://schemas.openxmlformats.org/drawingml/2006/main'}
                for color_scheme in root.findall('a:themeElements/a:clrScheme', ns):
                    for i, color in enumerate(color_scheme):
                        rgb = color.find('a:srgbClr', ns)
                        if rgb is not None:
                            theme_colors[i] = rgb.attrib['val']
    except Exception:
        logger.error("Erreur lors de l'extraction des couleurs du thème", exc_info=True)
    return theme_colors


def _find_columns_by_header(sheet, headers_wanted):
    """
    sheet : feuille openpyxl
    headers_wanted : dict {nom_logique: [liste de textes possibles dans la cellule]}
    retourne : dict {nom_logique: index_colonne (0‑based)} ou None si manquant
    """
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    indices = {}

    for logical_name, candidates in headers_wanted.items():
        idx = None
        for i, cell_value in enumerate(header_row):
            if cell_value is None:
                continue
            cell_text = str(cell_value).strip().lower()
            if cell_text in candidates:
                idx = i
                break
        if idx is None:
            logger.error("Colonne '%s' introuvable dans l'en‑tête : %s", logical_name, header_row)
            return None
        indices[logical_name] = idx

    return indices


def get_implantation_colors(file_path: str, sheet_name: str) -> dict:
    if not os.path.exists(file_path):
        logger.error("Fichier source introuvable pour get_implantation_colors : %s", file_path)
        return {}

    try:
        theme_colors = extract_theme_colors(file_path)
        workbook = load_workbook(filename=file_path, data_only=True)
    except Exception:
        logger.error("Erreur lors de l'ouverture du fichier source : %s", file_path, exc_info=True)
        return {}

    try:
        if sheet_name not in workbook.sheetnames:
            logger.error("Feuille source introuvable : %s dans %s", sheet_name, file_path)
            return {}

        sheet = workbook[sheet_name]

        # Recherche dynamique des colonnes
        col_indices = _find_columns_by_header(
            sheet,
            {
                "implantation": ["implantation"],
                "nom": ["nom"],
                "prenom": ["prénom", "prenom"],
            },
        )
        if col_indices is None:
            return {}

        idx_impl = col_indices["implantation"]
        idx_nom = col_indices["nom"]
        idx_prenom = col_indices["prenom"]

        data_colors = {}

        # on parcourt à partir de la 2e ligne (1 = en‑tête)
        for row in sheet.iter_rows(min_row=2):
            implantation = row[idx_impl].value
            nom = row[idx_nom].value
            prenom = row[idx_prenom].value

            if implantation is None or nom is None or prenom is None:
                continue

            key = (implantation, nom, prenom)

            cell_impl = row[idx_impl]
            if cell_impl.fill and cell_impl.fill.fill_type != "none":
                bg_color = cell_impl.fill.fgColor
                rvb_color = None

                if bg_color.type == "rgb":
                    rvb_color = hex_to_rvb(bg_color.rgb)
                elif bg_color.type == "theme":
                    hex_color = theme_colors.get(bg_color.theme)
                    rvb_color = hex_to_rvb(hex_color)

                # Ignorer les couleurs noires ou nulles
                if rvb_color and rvb_color != (0, 0, 0):
                    data_colors[key] = rvb_color

        return data_colors
    except Exception:
        logger.error("Erreur lors de l'extraction des couleurs d'implantation", exc_info=True)
        return {}
    finally:
        workbook.close()


def apply_colors_to_file2(file1_path: str, file1_sheet: str, file2_path: str, file2_sheet: str) -> None:
    if not os.path.exists(file2_path):
        logger.error("Fichier cible introuvable : %s", file2_path)
        return

    data_colors = get_implantation_colors(file1_path, file1_sheet)

    try:
        workbook = load_workbook(file2_path)
    except Exception:
        logger.error("Erreur lors de l'ouverture du fichier cible : %s", file2_path, exc_info=True)
        return

    try:
        if file2_sheet not in workbook.sheetnames:
            logger.error("Feuille cible introuvable : %s dans %s", file2_sheet, file2_path)
            return

        sheet = workbook[file2_sheet]

        # même logique : trouver les colonnes Implantation/Nom/Prénom dans le fichier 2
        col_indices = _find_columns_by_header(
            sheet,
            {
                "implantation": ["implantation"],
                "nom": ["nom"],
                "prenom": ["prénom", "prenom"],
            },
        )
        if col_indices is None:
            return

        idx_impl = col_indices["implantation"]
        idx_nom = col_indices["nom"]
        idx_prenom = col_indices["prenom"]

        for row in sheet.iter_rows(min_row=2):
            implantation = row[idx_impl].value
            nom = row[idx_nom].value
            prenom = row[idx_prenom].value

            if implantation is None or nom is None or prenom is None:
                continue

            key = (implantation, nom, prenom)
            rvb_color = data_colors.get(key)

            if rvb_color:
                fill = PatternFill(
                    start_color=f"{rvb_color[0]:02X}{rvb_color[1]:02X}{rvb_color[2]:02X}",
                    end_color=f"{rvb_color[0]:02X}{rvb_color[1]:02X}{rvb_color[2]:02X}",
                    fill_type="solid",
                )
                # applique la couleur sur toute la ligne (ou seulement les 3 colonnes si tu préfères)
                for cell in row:
                    cell.fill = fill

        workbook.save(file2_path)
    except Exception:
        logger.error("Erreur lors de l'application des couleurs au fichier cible", exc_info=True)
    finally:
        workbook.close()